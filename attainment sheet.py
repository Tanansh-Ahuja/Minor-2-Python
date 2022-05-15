import openpyxl as op
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side


def data(sheet, row, col, x, name='Times New Roman', bold=False, size=12, halign='left', valign='center', wrap=False):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    sheet.cell(row, col).value = x
    sheet.cell(row, col).font = Font(name=name, bold=bold, size=size)
    sheet.cell(row, col).border = thin_border
    sheet.cell(row, col).alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)


def attainment_theory_sheet(sheet):
    sheet.title = "Final CO-PO-PSO Attain_Theory"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    data(sheet, 1, 1, 'JAYPEE INSTITUTE OF INFORMATION TECHNOLOGY', bold=True)
    for i in range(1, 8):
        sheet.cell(1, i).border = thin_border

    for i in range(2, 7):
        sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        sheet.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)

    data(sheet, 2, 1, 'Academic Year')
    data(sheet, 2, 4, '2021-22(Odd Semester)', bold=True)
    data(sheet, 3, 1, 'Semester/Branch:')
    data(sheet, 3, 4, 'CSE', bold=True)
    data(sheet, 4, 1, 'NBA Code')
    data(sheet, 4, 4, 'Some code', bold=True)
    data(sheet, 5, 1, 'Course Name and Code: ')
    data(sheet, 5, 4, '18B11412CL113', bold=True)
    data(sheet, 6, 1, 'Course Coordinator(s): ')
    data(sheet, 6, 4, 'Course Coordinator Name', bold=True)
    for i in range(2, 7):
        for j in range(1, 7):
            sheet.cell(i, j).border = thin_border

    sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=12)
    data(sheet, 7, 1, 'Average CO-Attainment', bold=True, size=14, halign='center')
    for i in range(1, 13):
        sheet.cell(7, i).border = thin_border

    data(sheet, 8, 1, 'COs', bold=True, size=11, wrap=True, valign='center')
    data(sheet, 8, 2, 'T1', bold=True, size=11, wrap=True, valign='center')
    data(sheet, 8, 3, 'T2', bold=True, size=11, wrap=True, valign='center')
    data(sheet, 8, 4, 'End Term', bold=True, size=10, wrap=True, valign='center')
    data(sheet, 8, 5, 'T-AVG (Avg. of T1, T2 and T3)', bold=True, size=9, wrap=True, valign='center')

    average_co_attainment = [
        ['c113.1', [3, -1, 1], [2, -1, -1], 3],
        ['c113.2', [0, 0, 2], [-1, 3, -1], 3],
        ['c113.3', [-1, 2, 3], [-1, -1, 3], 3],
        ['c113.4', [-1, -1, 2], [-1, -1, -1], 3],
        ['c113.5', [-1, -1, 3], [-1, -1, -1], 3]
    ]
    ass_col = 6
    temp = 1
    for i in range(ass_col, ass_col + len(average_co_attainment[0][2])):
        data(sheet, 8, i, 'Assign-' + str(temp), bold=True, size=11, wrap=True, valign='center')
        temp += 1

    b = ass_col + len(average_co_attainment[0][2])
    data(sheet, 8, b, 'Assgn-AVG', size=9, bold=True, wrap=True, valign='center')

    data(sheet, 8, b + 1, '80% of Direct Assessment (60% T-AVG + 20% Assgn-AVG if Assgn component is used)', bold=True,
         size=9, wrap=True, valign='center')
    data(sheet, 8, b + 2, 'Student Feedback (Indirect Assessment)', bold=True, size=9, wrap=True, valign='center')
    data(sheet, 8, b + 3, 'Final (Direct + 20% Indirect)', bold=True, size=9, wrap=True, valign='center')

    for i in range(0, len(average_co_attainment)):
        total = 0
        counter = 0
        sheet.cell(9 + i, 1).value = average_co_attainment[i][0]

        if average_co_attainment[i][1][0] != -1:
            sheet.cell(9 + i, 2).value = average_co_attainment[i][1][0]
            total += average_co_attainment[i][1][0]
            counter += 1
        if average_co_attainment[i][1][1] != -1:
            sheet.cell(9 + i, 3).value = average_co_attainment[i][1][1]
            total += average_co_attainment[i][1][1]
            counter += 1
        if average_co_attainment[i][1][2] != -1:
            sheet.cell(9 + i, 4).value = average_co_attainment[i][1][2]
            total += average_co_attainment[i][1][2]
            counter += 1

        avg = round(total / counter, 1)
        sheet.cell(9 + i, 5).value = avg
        total2 = 0
        counter2 = 0
        for j in range(len(average_co_attainment[0][2])):
            if average_co_attainment[i][2][j] != -1:
                sheet.cell(9 + i, 6 + j).value = average_co_attainment[i][2][j]
                total2 += average_co_attainment[i][2][j]
                counter2 += 1
        j = len(average_co_attainment[0][2])
        if counter2 != 0:
            sheet.cell(9 + i, 6 + j).value = round(total2 / counter2, 1)

        var1 = sheet.cell(9 + i, 5).value
        var2 = sheet.cell(9 + i, 6 + j).value
        if var1 is None:
            var3 = 0.8 * var2
        elif var2 is None:
            var3 = 0.8 * var1
        else:
            var3 = 0.6 * var1 + 0.2 * var2
        sheet.cell(9 + i, 6 + j + 1).value = round(var3, 1)
        sheet.cell(9 + i, 6 + j + 2).value = average_co_attainment[i][3]
        var1 = sheet.cell(9 + i, 6 + j + 1).value
        var2 = sheet.cell(9 + i, 6 + j + 2).value
        if var1 is None:
            var1 = 0
        if var2 is None:
            var2 = 0
        sheet.cell(9 + i, 6 + j + 3).value = var1 + 0.2 * var2

    temp = 9
    while True:
        if sheet.cell(temp, 1).value is None:
            break
        for temp2 in range(1, 13):
            sheet.cell(temp, temp2).border = thin_border
        temp += 1

    start_row = temp + 1

    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=16)
    sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=2)

    data(sheet, start_row, 1, 'CO-PO-PSO Mapping', size=14, bold=True, halign='center')
    for temp in range(1, 17):
        sheet.cell(start_row, temp).border = thin_border

    table_heading_row = start_row + 1
    data(sheet, table_heading_row, 1, 'CO Attainments', size=10, bold=True, halign='center')
    sheet.cell(table_heading_row, 2).border = thin_border
    for temp in range(3, 15):
        data(sheet, table_heading_row, temp, 'PO{}'.format(str(temp - 2)), halign='center', bold=True)

    data(sheet, table_heading_row, 15, 'PS01', halign='center', bold=True)
    data(sheet, table_heading_row, 16, 'PS02', halign='center', bold=True)

    co_po_pso_mapping_table = [
        [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
        [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
        [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
        [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
        [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
    ]
    table_content = table_heading_row + 1
    table_end = table_content + len(average_co_attainment)
    avcoa = 0
    table1_end_index = 5 + len(average_co_attainment[0][2]) + 4
    for i in range(table_content, table_end):
        data(sheet, i, 1, average_co_attainment[avcoa][0], halign='center')
        data(sheet, i, 2, sheet.cell(9 + avcoa, table1_end_index).value, halign='center')
        avcoa += 1
        temp = 0

        for j in range(len(co_po_pso_mapping_table[0])):
            if co_po_pso_mapping_table[temp][j] != -1:
                data(sheet, i, j + 3, co_po_pso_mapping_table[temp][j])
            else:
                sheet.cell(i, j + 3).border = thin_border
        temp += 1

    table3_start = table_end + 1
    sheet.merge_cells(start_row=table3_start, end_row=table3_start, start_column=1, end_column=15)
    data(sheet, table3_start, 1, 'PO-PSO-Attainment', bold=True, size=14, halign='center')
    for i in range(0, 15):
        sheet.cell(table3_start, i + 1).border = thin_border
    data(sheet, table3_start + 1, 1, 'Course', bold=True, halign='center')

    for i in range(1, 13):
        data(sheet, table3_start + 1, i + 1, 'PO{}'.format(str(i)), bold=True, halign='center')
    data(sheet, table3_start + 1, 14, 'PSO1', bold=True, halign='center')
    data(sheet, table3_start + 1, 15, 'PSO2', bold=True, halign='center')

    data(sheet, table3_start + 2, 1, average_co_attainment[0][0][:-2], bold=True, halign='center')

    for i in range(3, 15):
        some_value = 0
        den = 0
        var1 = table_content
        var2 = var1
        while var2 < var1 + len(co_po_pso_mapping_table):

            if sheet.cell(var1, i).value is None:
                x = 0
            else:
                x = sheet.cell(var1, i).value
                den = den + x

            y = sheet.cell(var1, 2).value
            some_value += x * y
            var2 += 1
        if den != 0:
            data(sheet, table3_start + 2, i - 1, round(some_value / den, 1))
    for i in range(2, 16):
        sheet.cell(table3_start + 2, i).border = thin_border
    wb.save('./attainment sheet.xlsx')


# main
wb = op.Workbook('./attainment sheet.xlsx')
wb.save('./attainment sheet.xlsx')
wb = op.load_workbook('./attainment sheet.xlsx')
sheet = wb.active
attainment_theory_sheet(sheet)

# start on lab sheet
wb.create_sheet('Final CO-PO-PSO Attainment_Lab')
sheet = wb['Final CO-PO-PSO Attainment_Lab']
sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
data(sheet, 1, 1, 'JAYPEE INSTITUTE OF INFORMATION TECHNOLOGY', bold=True)
for i in range(1, 8):
    sheet.cell(1, i).border = thin_border

for i in range(2, 7):
    sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    sheet.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
# DATA TABLE
data(sheet, 2, 1, 'Academic Year')
data(sheet, 2, 4, '2021-22(Odd Semester)', bold=True)
data(sheet, 3, 1, 'Semester/Branch:')
data(sheet, 3, 4, 'CSE', bold=True)
data(sheet, 4, 1, 'NBA Code')
data(sheet, 4, 4, 'Some code', bold=True)
data(sheet, 5, 1, 'Course Name and Code: ')
data(sheet, 5, 4, '18B11412CL113', bold=True)
data(sheet, 6, 1, 'Course Coordinator(s): ')
data(sheet, 6, 4, 'Course Coordinator Name', bold=True)
for i in range(2, 7):
    for j in range(1, 7):
        sheet.cell(i, j).border = thin_border



average_co_attainment = [
    # ['course code',eval,labtest,project,student feedback]
    ['c113.1', [2, -1], [-1, -1], 3, 3],
    ['c113.2', [-1, 3], [2, -1], 3, 3],
    ['c113.3', [-1, -1], [-1, -1], 3, 3],
    ['c113.4', [-1, 3], [-1, 3], 3, 3],
    ['c113.5', [-1, -1], [-1, 2], 3, 3]
]

# TABLE 1
data(sheet, 8, 1, 'COs', bold=True, size=11, wrap=True, valign='center')
for i in range(len(average_co_attainment[0][1])):
    data(sheet, 8, 2 + i, 'Eval-{}'.format(str(i + 1)), bold=True, wrap=True, valign='center')
lab_text_index = 2 + len(average_co_attainment[0][1])
for i in range(len(average_co_attainment[0][2])):
    data(sheet, 8, lab_text_index + i, 'Lab test-{}'.format(str(i + 1)), bold=True, wrap=True, valign='center',size=10)
project_index = lab_text_index + len(average_co_attainment[0][2])
data(sheet, 8, project_index, 'Project [15]', bold=True, wrap=True, valign='center',size=10)
data(sheet, 8, project_index + 1, 'Direct Attainment', bold=True, wrap=True, valign='center',size=10)
data(sheet, 8, project_index + 2, 'Student Feedback (Indirect Assessment)', bold=True, wrap=True, valign='center',size=9)
data(sheet, 8, project_index + 3, 'Final (80% Direct + 20% Indirect)', bold=True, wrap=True, valign='center',size=9)

# MIDDLE OF DATA TABLE  AND TABLE 1
sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=project_index+3)
data(sheet, 7, 1, 'Average CO-Attainment', bold=True, size=14, halign='center')
for i in range(1, project_index+3+1):
    sheet.cell(7, i).border = thin_border

# INSERTING VALUES IN TABLE 1
table2_start = 9
for i in range(len(average_co_attainment)):
    # ROW WISE
    data(sheet, i + table2_start, 1, average_co_attainment[i][0], bold=True, halign='center')
    #EVAL
    for j in range(len(average_co_attainment[0][1])):
        if average_co_attainment[i][1][j] != -1:
            data(sheet, i + table2_start, 2 + j, average_co_attainment[i][1][j], halign='center')
        else:
            sheet.cell(i + table2_start, 2 + j).border = thin_border
    # LAB TEST
    for j in range(len(average_co_attainment[0][2])):
        if average_co_attainment[i][2][j] != -1:
            data(sheet, i + table2_start, lab_text_index + j, average_co_attainment[i][2][j], halign='center')
        else:
            sheet.cell(i + table2_start, lab_text_index + j).border = thin_border
    # PROJECT
    if average_co_attainment[i][3] != -1:
        data(sheet, i + table2_start, project_index, average_co_attainment[i][3], halign='center')
    else:
        sheet.cell(i + table2_start, project_index).border = thin_border
    # STUDENT FEEDBACK
    if average_co_attainment[i][4] != -1:
        data(sheet, i + table2_start, project_index + 2, average_co_attainment[i][4], halign='center')
    else:
        sheet.cell(i + table2_start, project_index + 2).border = thin_border

    # calculation
    total = 0
    counter = 0
    for j in range(2, project_index + 1):
        if sheet.cell(i + table2_start, j).value is not None:
            total = total + sheet.cell(i + table2_start, j).value
            counter += 1

    data(sheet, i + table2_start, project_index + 1, round(total / counter, 1), halign='center')
    data(sheet, i + table2_start, project_index + 3,
         round(0.8 * sheet.cell(i + table2_start, project_index + 1).value + 0.2 * sheet.cell(i + table2_start,
                                                                                              project_index + 2).value,
               1),
         halign='center')

# TABLE 1 ENDS

# TABLE 2 BEGINS
start_row = len(average_co_attainment) + 9 + 1
sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=16)
sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=2)

data(sheet, start_row, 1, 'CO-PO-PSO Mapping', size=14, bold=True, halign='center')
for temp in range(1, 17):
    sheet.cell(start_row, temp).border = thin_border

table_heading_row = start_row + 1
data(sheet, table_heading_row, 1, 'CO Attainments', size=10, bold=True, halign='center')
sheet.cell(table_heading_row, 2).border = thin_border
for temp in range(3, 15):
    data(sheet, table_heading_row, temp, 'PO{}'.format(str(temp - 2)), halign='center', bold=True)

data(sheet, table_heading_row, 15, 'PS01', halign='center', bold=True)
data(sheet, table_heading_row, 16, 'PS02', halign='center', bold=True)

co_po_pso_mapping_table = [
    [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
    [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
    [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
    [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
    [2, 2, 2, -1, 2, 1, -1, 1, 1, 1, 1, 2, 2, 2],
]
# INSERTING VALUES IN TABLE 2
table_content = table_heading_row + 1
table_end = table_content + len(average_co_attainment)
avcoa = 0
table1_end_index = project_index + 3
for i in range(table_content, table_end):
    data(sheet, i, 1, average_co_attainment[avcoa][0], halign='center')
    data(sheet, i, 2, sheet.cell(9 + avcoa, table1_end_index).value, halign='center')
    avcoa += 1
    temp = 0

    for j in range(len(co_po_pso_mapping_table[0])):
        if co_po_pso_mapping_table[temp][j] != -1:
            data(sheet, i, j + 3, co_po_pso_mapping_table[temp][j])
        else:
            sheet.cell(i, j + 3).border = thin_border
    temp += 1

# TABLE 2 ENDS

# TABLE 3 BEGINS
table3_start = table_end + 1
sheet.merge_cells(start_row=table3_start, end_row=table3_start, start_column=1, end_column=15)
data(sheet, table3_start, 1, 'PO-PSO-Attainment', bold=True, size=14, halign='center')
for i in range(0, 15):
    sheet.cell(table3_start, i + 1).border = thin_border
data(sheet, table3_start + 1, 1, 'Course', bold=True, halign='center')

for i in range(1, 13):
    data(sheet, table3_start + 1, i + 1, 'PO{}'.format(str(i)), bold=True, halign='center')
data(sheet, table3_start + 1, 14, 'PSO1', bold=True, halign='center')
data(sheet, table3_start + 1, 15, 'PSO2', bold=True, halign='center')

# INSERTING VALUES IN TABLE 3
data(sheet, table3_start + 2, 1, average_co_attainment[0][0][:-2], bold=True, halign='center')

for i in range(3, 15):
    some_value = 0
    den = 0
    var1 = table_content
    var2 = var1
    while var2 < var1 + len(co_po_pso_mapping_table):

        if sheet.cell(var2, i).value is None:
            x = 0
        else:
            x = sheet.cell(var2, i).value
            den = den + x

        y = sheet.cell(var2, 2).value
        some_value += x * y
        var2 += 1
    if den != 0:
        data(sheet, table3_start + 2, i - 1, round(some_value / den, 1))
for i in range(2, 16):
    sheet.cell(table3_start + 2, i).border = thin_border

wb.save('./attainment sheet.xlsx')
